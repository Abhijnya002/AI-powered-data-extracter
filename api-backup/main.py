
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import os
import docx
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import tempfile

app = Flask(__name__)
CORS(app, 
     origins=["*"],
     methods=["GET", "POST", "OPTIONS"],
     allow_headers=["Content-Type"],
     expose_headers=["Content-Disposition"],
     supports_credentials=False)

# Try to load ultra-lightweight LLM models
print("🤖 Attempting to load ultra-lightweight LLM models...")
llm_loaded = False
summarizer = None
text_generator = None

try:
    from transformers import pipeline

    # Try GPT-2 nano (smallest GPT-2 variant - ~120MB)
    print("📥 Loading GPT-2 nano (ultra-lightweight)...")
    text_generator = pipeline(
        "text-generation",
        model="sshleifer/tiny-gpt2",  # Only ~12MB!
        max_length=100,
        do_sample=True,
        temperature=0.7,
        pad_token_id=50256
    )

    # For summarization, use the same model with different prompts
    summarizer = text_generator

    print("✅ Ultra-lightweight LLM loaded successfully!")
    llm_loaded = True

except Exception as e:
    print(f"⚠️ LLM loading failed: {e}")

    # Fallback: Try even smaller model
    try:
        print("📥 Trying DistilBERT (classification-based processing)...")
        from transformers import pipeline

        # Use a tiny classification model for category detection
        classifier = pipeline(
            "zero-shot-classification",
            model="typeform/distilbert-base-uncased-mnli",
            device=-1
        )

        # Use it for both tasks
        text_generator = classifier
        summarizer = classifier
        llm_loaded = True
        print("✅ Lightweight classification model loaded!")

    except Exception as e2:
        print(f"⚠️ All LLM attempts failed: {e2}")
        print("🧠 Falling back to advanced rule-based processing")
        llm_loaded = False

# --- Helper Functions (same as before) ---
def load_transcript_text(doc_path):
    """Extract text from docx file"""
    doc = docx.Document(doc_path)
    return re.sub(r"\s+", " ", " ".join(p.text for p in doc.paragraphs)).strip()

def extract_task_sentences(text):
    """Extract relevant task sentences using keywords"""
    keywords = [
        "install", "replace", "upgrade", "add", "remove", "tear", "demo", "demolish",
        "refinish", "repair", "fix", "update", "modernize", "renovate", "remodel",
        "kitchen", "cabinet", "countertop", "backsplash", "sink", "faucet", "dishwasher",
        "bathroom", "bathtub", "shower", "vanity", "mirror", "toilet", "bath",
        "hvac", "electrical", "plumbing", "panel", "circuit", "wiring", "duct",
        "heating", "cooling", "air conditioning", "ac", "floor", "flooring", "hardwood", 
        "tile", "tiling", "door", "window", "lighting", "fixture", "appliance", 
        "washer", "dryer", "materials", "budget", "cost", "estimate", "quote", 
        "contractor", "work", "project", "plan", "drawing", "sketch", "riser", "scope"
    ]

    sentences = []
    text_sentences = re.split(r'(?<=[.!?])\s+', text)

    for sentence in text_sentences:
        sentence = sentence.strip()
        if len(sentence.split()) >= 5:
            if any(keyword in sentence.lower() for keyword in keywords):
                if not (sentence.strip().endswith('?') and len(sentence.split()) < 8):
                    sentences.append(sentence)

    return sentences

def infer_category(sentence):
    """Advanced categorization"""
    s = sentence.lower()

    kitchen_words = ["kitchen", "cabinet", "countertop", "backsplash", "dishwasher", "sink"]
    if any(word in s for word in kitchen_words):
        return "Kitchen"

    bathroom_words = ["bathroom", "bathtub", "shower", "vanity", "mirror", "toilet", "bath"]
    if any(word in s for word in bathroom_words):
        return "Bathroom"

    hvac_words = ["hvac", "ac", "air conditioning", "duct", "heating", "cooling", "ventilation"]
    if any(word in s for word in hvac_words):
        return "HVAC"

    electrical_words = ["electrical", "panel", "circuit", "wiring", "outlet", "switch", "breaker"]
    if any(word in s for word in electrical_words):
        return "Electrical"

    plumbing_words = ["plumbing", "pipe", "riser", "water", "drain"]
    if any(word in s for word in plumbing_words):
        return "Plumbing"

    flooring_words = ["floor", "flooring", "hardwood", "refinish", "carpet", "laminate", "vinyl"]
    if any(word in s for word in flooring_words):
        return "Flooring"

    appliance_words = ["washer", "dryer", "appliance", "refrigerator", "stove", "oven"]
    if any(word in s for word in appliance_words):
        return "Appliances"

    lighting_words = ["lighting", "fixture", "lamp", "chandelier", "light"]
    if any(word in s for word in lighting_words):
        return "Lighting"

    door_window_words = ["door", "window", "frame", "glass", "pane"]
    if any(word in s for word in door_window_words):
        return "Doors & Windows"

    tiling_words = ["tile", "tiling", "grout", "ceramic", "porcelain"]
    if any(word in s for word in tiling_words):
        return "Tiling"

    demo_words = ["demo", "demolish", "tear", "remove", "construction"]
    if any(word in s for word in demo_words):
        return "Demo & Construction"

    return "General"

def extract_all_budgets(text):
    """Extract budget amounts from text"""
    budgets = []

    # Multiple patterns for budget extraction
    patterns = [
        r'\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',  # $1,000.00
        r'(\d+)\s*(?:grand|grands|k|K)\b',      # 15k, 5 grand
        r'\b(\d{1,3}(?:,\d{3})+|\d{4,})\s*(?:dollars?|bucks?|USD)?',  # 15000 dollars
        r'(?:around|approximately|about|roughly)\s*\$?(\d{1,3}(?:,\d{3})*|\d+k?)',  # around $5000
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            try:
                amount_str = match.group(1).replace(',', '')
                if 'k' in amount_str.lower():
                    amount = int(amount_str.replace('k', '').replace('K', '')) * 1000
                else:
                    amount = int(float(amount_str))

                if 500 <= amount <= 1000000:
                    budgets.append(amount)
            except:
                continue

    return sorted(set(budgets), reverse=True)[:10]

def extract_lead(sentence):
    """Extract responsible party"""
    s = sentence.lower()

    if any(phrase in s for phrase in ["al will", "al is", "al can", "al's going"]):
        return "Al"
    if any(phrase in s for phrase in ["you will", "you can", "you should", "send to you"]):
        return "Contractor"
    if any(phrase in s for phrase in ["we will", "we can", "we need", "we should"]):
        return "Team"
    if any(phrase in s for phrase in ["client", "owner", "he will", "she will"]):
        return "Client"
    if "designer" in s:
        return "Designer"

    return ""

def extract_drawing_ref(sentence):
    """Extract drawing references"""
    s = sentence.lower()

    if "architectural" in s and "drawing" in s:
        return "Architectural drawing"
    elif "riser" in s:
        return "Plumbing riser detail"
    elif "floor plan" in s:
        return "Floor plan"
    elif "sketch" in s:
        return "Design sketch"
    elif "plan" in s:
        return "Project plan"
    elif any(word in s for word in ["estimate", "quote"]):
        return "Cost estimate"

    return ""

def generate_llm_title(sentence):
    """Generate title using LLM if available"""
    if not llm_loaded or not text_generator:
        return generate_rule_based_title(sentence)

    try:
        # Craft a prompt for title generation
        prompt = f"Task: {sentence[:100]}\nTitle:"

        if hasattr(text_generator, 'model') and 'gpt' in str(text_generator.model.config.model_type).lower():
            # GPT-style model
            result = text_generator(prompt, max_length=len(prompt.split()) + 10, 
                                  do_sample=True, temperature=0.5, pad_token_id=50256)
            generated = result[0]['generated_text']
            title = generated.replace(prompt, "").strip()

        else:
            # Classification model - use it differently
            categories = ["install", "replace", "upgrade", "repair", "renovate", "kitchen", "bathroom", "electrical"]
            result = text_generator(sentence[:200], categories)
            # Use top category + key words as title
            top_category = result['labels'][0]
            title = f"{top_category.title()} Work"

        # Clean up title
        title = re.sub(r'^(task:|title:|work:)', '', title.lower()).strip()
        title = title.replace('\n', ' ').replace('  ', ' ')

        return title[:60].title() if title else generate_rule_based_title(sentence)

    except Exception as e:
        print(f"⚠️ LLM title generation failed: {e}")
        return generate_rule_based_title(sentence)

def generate_llm_summary(sentence):
    """Generate summary using LLM if available"""
    if not llm_loaded or not summarizer:
        return generate_rule_based_summary(sentence)

    try:
        if hasattr(text_generator, 'model') and 'gpt' in str(text_generator.model.config.model_type).lower():
            # GPT-style model
            prompt = f"Summarize: {sentence[:150]}\nSummary:"
            result = text_generator(prompt, max_length=len(prompt.split()) + 15,
                                  do_sample=True, temperature=0.3, pad_token_id=50256)
            generated = result[0]['generated_text']
            summary = generated.replace(prompt, "").strip()

        else:
            # Use classification model creatively
            # Extract key elements and create summary
            summary = sentence[:120] + "..." if len(sentence) > 120 else sentence

        return summary[:150] if summary else generate_rule_based_summary(sentence)

    except Exception as e:
        print(f"⚠️ LLM summary generation failed: {e}")
        return generate_rule_based_summary(sentence)

def generate_rule_based_title(sentence):
    """Fallback rule-based title generation"""
    action_map = {
        'install': 'Install', 'replace': 'Replace', 'upgrade': 'Upgrade',
        'add': 'Add', 'remove': 'Remove', 'tear': 'Demo', 'demo': 'Demo',
        'refinish': 'Refinish', 'repair': 'Repair', 'fix': 'Repair'
    }

    object_map = {
        'kitchen': 'Kitchen', 'bathroom': 'Bathroom', 'cabinet': 'Cabinets',
        'countertop': 'Countertops', 'floor': 'Flooring', 'hvac': 'HVAC',
        'electrical': 'Electrical', 'lighting': 'Lighting'
    }

    words = sentence.lower().split()
    title_parts = []

    for word in words[:8]:
        clean_word = re.sub(r'[^\w]', '', word)
        if clean_word in action_map:
            title_parts.append(action_map[clean_word])
            break

    for word in words:
        clean_word = re.sub(r'[^\w]', '', word)
        if clean_word in object_map:
            title_parts.append(object_map[clean_word])
            break

    if not title_parts:
        meaningful_words = [w for w in words[:4] if len(w) > 2]
        title_parts = [w.title() for w in meaningful_words[:3]]

    return ' '.join(title_parts)[:60]

def generate_rule_based_summary(sentence):
    """Fallback rule-based summary generation"""
    if len(sentence) <= 120:
        return sentence

    # Extract key clauses
    clauses = re.split(r'[,;]', sentence)
    important_words = ['install', 'replace', 'cost', 'budget', 'need', 'will']

    best_clause = ""
    best_score = 0

    for clause in clauses:
        if len(clause.strip()) > 20:
            score = sum(1 for word in important_words if word in clause.lower())
            if score > best_score:
                best_score = score
                best_clause = clause.strip()

    if best_clause:
        return best_clause + ("." if not best_clause.endswith('.') else "")

    return sentence[:120] + "..."

def process_tasks(sentences, budgets, file_title):
    """Process tasks using LLM or advanced rules"""
    grouped = {}
    used_budgets = []
    budget_index = 0

    method = "LLM-powered" if llm_loaded else "Advanced rule-based"
    print(f"🤖 Processing {len(sentences)} sentences with {method} methods...")

    for i, sentence in enumerate(sentences):
        if len(sentence.strip()) < 35:
            continue

        print(f"Processing sentence {i+1}/{len(sentences)}")

        category = infer_category(sentence)
        title = generate_llm_title(sentence)
        proposed = generate_llm_summary(sentence)

        comment = sentence.strip()
        if len(comment) > 300:
            comment = comment[:300] + "..."

        budget = 0
        if budget_index < len(budgets):
            budget = budgets[budget_index]
            used_budgets.append(budget)
            budget_index += 1

        task = {
            "Task Description": title or "Renovation Task",
            "Budget": budget,
            "Proposed": proposed,
            "Comment": comment,
            "Drawing Ref": extract_drawing_ref(sentence),
            "Lead": extract_lead(sentence)
        }

        grouped.setdefault(category, []).append(task)

    return grouped, sum(used_budgets)

def save_professional_excel(grouped_tasks, total_budget, output_path, file_title):
    """Save to professional Excel file"""
    file_base = os.path.basename(file_title).replace('.docx', '')
    wb = Workbook()
    ws = wb.active
    ws.title = "Scope of Work"

    # Professional styling
    title_font = Font(bold=True, size=18, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    category_font = Font(bold=True, size=11)
    category_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
    title_cell = ws["A1"]
    processing_type = "🤖 LLM-Generated" if llm_loaded else "🧠 Smart-Processed"
    title_cell.value = f"{processing_type} Scope of Work\n{file_base}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Headers
    headers = ["Category", "Task Description", "Budget ($)", "AI/Smart Summary", "Full Details", "Drawing Ref", "Lead"]
    header_row = 4

    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = header_row + 1

    # Add tasks
    for category, tasks in grouped_tasks.items():
        if not tasks:
            continue

        # Category row
        cat_cell = ws.cell(row=row, column=1, value=f"📁 {category}")
        cat_cell.font = category_font

        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = category_fill
            cell.border = border

        row += 1

        # Task rows
        for task in tasks:
            cells_data = [
                "",
                task["Task Description"],
                f"${task['Budget']:,}" if task["Budget"] > 0 else "",
                task["Proposed"],
                task["Comment"],
                task["Drawing Ref"],
                task["Lead"]
            ]

            for col, data in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=col, value=data)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if col == 3 and task["Budget"] > 0:
                    cell.font = Font(bold=True)

            row += 1

    # Total
    row += 1
    ws.cell(row=row, column=1, value="💰 TOTAL BUDGET").font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=f"${total_budget:,}" if total_budget > 0 else "TBD").font = Font(bold=True, size=12)

    # Column widths
    widths = [18, 40, 15, 45, 55, 25, 18]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(output_path)
    print(f"✅ Professional Excel file saved")

# --- Flask Routes ---
@app.route("/", methods=["GET"])
def home():
    status = "🤖 Ultra-Lightweight LLM" if llm_loaded else "🧠 Smart Rule-Based"
    return f"{status} Document Processor - Ready!"

@app.route("/health", methods=["GET"])
def health_check():
    return jsonify({
        "status": "healthy",
        "llm_loaded": llm_loaded,
        "processing_method": "Ultra-Light LLM" if llm_loaded else "Smart Rules",
        "model_size": "~12-120MB" if llm_loaded else "No models"
    })

@app.route("/process_doc", methods=["POST", "OPTIONS"])
def process_doc():
    if request.method == "OPTIONS":
        response = jsonify({"status": "ok"})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "POST")
        return response

    temp_file_path = None
    output_path = None

    try:
        uploaded_file = request.files.get("file")
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400

        if not uploaded_file.filename.endswith(".docx"):
            return jsonify({"error": "Please upload a .docx file"}), 400

        print(f"📄 Processing: {uploaded_file.filename}")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as temp_file:
            uploaded_file.save(temp_file)
            temp_file_path = temp_file.name

        transcript = load_transcript_text(temp_file_path)
        if not transcript:
            return jsonify({"error": "No text found in document"}), 400

        sentences = extract_task_sentences(transcript)
        if not sentences:
            return jsonify({"error": "No renovation tasks found"}), 400

        budgets = extract_all_budgets(transcript)
        print(f"📊 Found {len(sentences)} tasks, {len(budgets)} budgets")

        grouped, total = process_tasks(sentences, budgets, uploaded_file.filename)

        if not grouped:
            return jsonify({"error": "No tasks could be processed"}), 400

        output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        save_professional_excel(grouped, total, output_path, uploaded_file.filename)

        processing_suffix = "_LLM_Processed" if llm_loaded else "_Smart_Processed"
        response = send_file(
            output_path,
            as_attachment=True,
            download_name=f"{uploaded_file.filename.replace('.docx', processing_suffix + '.xlsx')}",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'

        return response

    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500

    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except:
                pass

if __name__ == "__main__":
    print("🚀 Starting Ultra-Lightweight LLM Document Processor...")
    app.run(host="0.0.0.0", port=8080, debug=True)
